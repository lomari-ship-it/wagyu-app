#!/usr/bin/env python3
"""
Generate filled Book2.xlsx birth notification from calf data passed as JSON.
Usage: python3 generate_book.py '<json_data>' > output.xlsx
"""
import sys
import json
import copy
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def generate(calves_json, template_path='/mnt/user-data/uploads/Book2.xlsx'):
    calves = json.loads(calves_json)
    wb = load_workbook(template_path)
    ws = wb['birth notification']

    # Row 4 is the template row with formulas - we fill from row 4 onwards
    START_ROW = 4

    for i, calf in enumerate(calves):
        row = START_ROW + i

        # If beyond row 4, copy formula structure from row 4
        if row > START_ROW:
            for col_letter in ['H', 'J', 'L', 'N', 'Q', 'AD', 'AG', 'AI']:
                src = ws[f'{col_letter}{START_ROW}']
                tgt = ws[f'{col_letter}{row}']
                # Replace row number in formula
                if src.value and isinstance(src.value, str):
                    tgt.value = src.value.replace(str(START_ROW), str(row))

        # Parse birth date
        dob = calf.get('birth_date', '')
        try:
            dt = datetime.strptime(dob, '%Y-%m-%d')
            day, month, year = dt.strftime('%d'), dt.strftime('%m'), dt.strftime('%Y')
        except:
            day, month, year = '', '', ''

        # Parse identity number - strip prefix (e.g. "26-0001JFW" -> name part varies)
        identity = calf.get('identity_number', '')

        # C = animal name (identity without breed prefix, e.g. "0001JFW" or full if no prefix)
        ws[f'C{row}'] = identity

        # E/F/G = D/M/Y of birth
        ws[f'E{row}'] = day
        ws[f'F{row}'] = month
        ws[f'G{row}'] = year

        # I = Sex text (Female/Male) - XLOOKUP in H reads from I
        sex = calf.get('sex', '')
        ws[f'I{row}'] = sex

        # K = Calf details text (Single/Twin/Multiple) - XLOOKUP in J reads from K
        calf_details = calf.get('calf_details', 'Single')
        ws[f'K{row}'] = calf_details

        # R = Calf ID number (HDM part = full identity)
        ws[f'R{row}'] = identity

        # U = Sire ID
        ws[f'U{row}'] = calf.get('father_id', '')

        # X = Dam ID
        ws[f'X{row}'] = calf.get('mother_id', '')

        # AH = Birth mass
        birth_mass = calf.get('birth_mass')
        if birth_mass:
            ws[f'AH{row}'] = float(birth_mass)

        # AJ = Colour text - XLOOKUP in AI reads from AJ
        ws[f'AJ{row}'] = calf.get('color', '')

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

if __name__ == '__main__':
    data = sys.argv[1] if len(sys.argv) > 1 else '[]'
    result = generate(data)
    sys.stdout.buffer.write(result)
